import csv
import requests
import os
import shutil
import sys
import getopt
import xlwt
from packages import ETD
from lxml import etree
from langcodes import *
from openpyxl import load_workbook
from datetime import datetime
import re

import envconfig

################################################################################
## ETD Migration Script                                                       ##
################################################################################
## This Script does the following:                                            ##
##  1.  Reads through a reference input CSV (generated by buildSheet.py)      ##
##  2.  Retrieve Alma Bibliographic data (via API)                            ##
##  3.  Retrieve ProQuest data (stored in ETD Storage)                        ##
##  4.  Deposit copies of each ETD's PDF file into a designated directory     ##
##  5.  Generate an XLS merging the above data for BePress consumption        ##
################################################################################

degree_map = {
    'D.A.':'Doctor of Arts (DA)',
    'Dr.P.H.':'Doctor of Public Health (DrPH)',
    'Ph.D.':'Doctor of Philosophy (PhD)',
    'Psy.D.':'Doctor of Psychology (PsyD)',
    'M.A.':'Master of Arts (MA)',
    'M.S.':'Master of Science (MS)',
    '':''
}

document_map = {
    'D.A.':'dissertation',
    'Dr.P.H.':'dissertation',
    'Ph.D.':'dissertation',
    'Psy.D.':'dissertation',
    'M.A.':'master\'s thesis',
    'M.S.':'master\'s thesis',
    '':''
}

def main(argv):
    print('Initiating...');
    
    #setup
    input_file = ''
    output_file = ''
    optout_file = ''
    takedown_file = ''
    disciplines_file = ''
    
    # script consumes user-provided values
    # -i flags the input file name
    # -o flags the output file name
    # -u flags the optout file name (optional)
    # -t flags the takedown file name (optional)
    # -d flags the discipline mapping file name (optional)
    
    opts, args = getopt.getopt(argv,'hi:o:u:t:d:')
    for opt, arg in opts:
        if opt == '-h':
            print ('migration.py -i <input-file.csv> -o <output-file.xls> [-u <optout-file.xlsx>] [-t <takedown-file.xlsx>] [-d <discipline-mapping-file.xlsx]')
            sys.exit()
        elif opt in ['-i']:
            input_file = arg
        elif opt in ['-o']:
            output_file = arg
        elif opt in ['-u']:
            optout_file = arg
        elif opt in ['-t']:
            takedown_file = arg
        elif opt in ['-d']:
            disciplines_file = arg
    #continue only if input and output arguments are not empty
    if not input_file:
        print('Input File not set')
    elif not output_file:
        print('Output File not set')
    else:
        print('Gathering data...')
        
        #create normal output workbook and output workbook for items that require manual attention
        output_wb = xlwt.Workbook()
        output_ws = output_wb.add_sheet(' ')
        
        manual_output_wb = xlwt.Workbook()
        manual_output_ws = manual_output_wb.add_sheet(' ')
        
        headers = [
            {'input':'title','output':'title'},
            {'input':'url','output':'fulltext_url'},
            {'input':'keyword,650','output':'keywords'},
            {'input':'para','output':'abstract'},
            {'input':'First-Name','output':'author1_fname'},
            {'input':'Middle-Name','output':'author1_mname'},
            {'input':'Last-Name','output':'author1_lname'},
            {'input':'Author-Suffix','output':'author1_suffix'},
            {'input':'ProQuest-Email','output':'author1_email'},
            {'input':'inst_name','output':'author1_institution'},
            {'input':'advisor','output':'advisor1'},
            {'input':'advisor2','output':'advisor2'},
            {'input':'advisor3','output':'advisor3'},
            {'input':'mms_id','output':'alma_mms_id'},
            {'input':'categorization','output':'disciplines'},
            #{'input':'050','output':'call_number'},
            #{'input':'338','output':'carrier_type'},
            {'input':'comments','output':'comments'},
            {'input':'committee_members','output':'committee_members'},
            #{'input':'336','output':'content_type'},
            {'input':'degree','output':'degree_name'},
            {'input':'inst_contact','output':'department'},
            {'input':'',    'output':'distribution_license'},               # NO INCOMING VALUE
            {'input':'document_type','output':'document_type'},
            {'input':'',    'output':'doi'},                                # NO INCOMING VALUE
            {'input':'active embargo','output':'embargo_date'},
            #{'input':'655','output':'genre_form'},
            {'input':'language','output':'language'},
            #{'input':'337','output':'media_type'},
            {'input':'',    'output':'oa_licenses'},                        # NO INCOMING VALUE
            {'input':'',    'output':'orcid'},                              # NO INCOMING VALUE
            {'input':'300','output':'phys_desc'},
            {'input':'date_of_publication','output':'publication_date'},
            {'input':'',    'output':'season'},                             # NO INCOMING VALUE
            {'input':'',    'output':'rights_statements'},                  # NO INCOMING VALUE
            #{'input':'650','output':'subjects'}
            ]
        
        #worksheet header row
        for column,header in enumerate(headers):
            output_ws.write(0,column,header['output'])
            manual_output_ws.write(0,column,header['output'])
        
        # read optout file
        optout_entries = set()
        
        if optout_file != '':
            optout_sheet = load_workbook(filename = optout_file, read_only = True)['in']
        
            print('Reading Optout File...')
            
            for i, row in enumerate(optout_sheet):
                if i > 0:
                    optout_entries.add(row[0].value)
            
        print('Optout set: ', optout_entries)
        
        # read takedown file
        takedown_entries = {}
        
        if takedown_file != '':
            takedown_sheet = load_workbook(filename = takedown_file, read_only = True)['catalog_embargo_issues']
        
            print('Reading ProQuest Takedown File...')
            
            for i, row in enumerate(takedown_sheet):
                if i > 0 and row[1].value:
                    # 'mms_id' column value as key, date extrapolated from 'End date' column value as value
                    mms_id = row[2].value
                    end_date = row[1].value
                    
                    if ';' in end_date:
                        date_string = end_date.split(';')[0]
                        takedown_entries[mms_id] = date_string
                    else:
                        takedown_entries[mms_id] = ''
        
        print('Takedown set: ', takedown_entries.keys())
        
        # read discipline mapping file
        disciplines = dict()
        
        if disciplines_file != '':
            disciplines_sheet = load_workbook(filename = disciplines_file, read_only = True)['proquest disciplines']
            
            for i, row in enumerate(disciplines_sheet):
                if i > 0:
                    key = row[0].value.split(':')[0].lstrip()
                    value = row[1].value or ''
                    
                    value_split = value.split(': ')
                    
                    if re.search(value, 'map individually', re.IGNORECASE):
                        disciplines[key] = '**' + key + '**'
                    else:
                        disciplines[key] = value_split[len(value_split) - 1]
        
        # read input file
        print('Parsing Input File...')
        with open(input_file, newline='', encoding="utf-8") as input_csv:  ### sometimes 'cp1252', sometimes 'utf-8'
            input_reader = csv.reader(input_csv)
            input_headers = next(input_reader)
            
            row_count = 0
            manual_output_row_count = 0
            
            for row_array in input_reader:
                mms_id          = row_array[0]
                etd_id          = row_array[1]
                xml_id          = row_array[2]
                year            = row_array[3]
                active_embargo  = row_array[4]
                
                # apply takedown end date, if applicable
                if mms_id in takedown_entries.keys():
                    active_embargo = takedown_entries[mms_id]
                
                print(' '.join([mms_id, etd_id, year, active_embargo]))
                
                # retrieve records and flatten
                bib_record = get_bib(mms_id)
                etd_record = get_etd(etd_id, xml_id, year)
                
                data = flatten_data(headers, disciplines, bib_record, etd_record, mms_id, etd_id, xml_id, year, active_embargo, mms_id in optout_entries, mms_id in takedown_entries.keys())
                
                # write entry into manual_output_ws if condition is triggered
                if data['disciplines'].find('**') > -1:
                    for column,header in enumerate(headers):
                        manual_output_ws.write(manual_output_row_count+1,column,data[header['output']])
                        print('\t','MAP DISCIPLINE MANUALLY')
                    manual_output_row_count += 1
                # write entry into output_ws
                else:
                    for column,header in enumerate(headers):
                        output_ws.write(row_count+1,column,data[header['output']])
                    row_count += 1
        #save output file
        output_wb.save(output_file)
        
        #save manual-input output file
        manual_output_file = output_file[0:output_file.rfind('.')] + '.man.' + output_file[output_file.rfind('.')+1:len(output_file)]
        manual_output_wb.save(manual_output_file)

#combines data from bib record and etd record to produce a flat object that contains the necessary field values
def flatten_data(headers, disciplines, bib_record, etd_record, mms_id, etd_id, xml_id, year, active_embargo, opted_out, taken_down):
    data = {}
    
    #this fills the row with an error report if the associated input data is corrupt
    if etd_record is None or bib_record is None:
        for i,header in enumerate(headers):
            if i == 0:
                if bib_record == None:
                    data[header['output']] = 'Bib Record Not Retrieved: ' + mms_id
                else:
                    data[header['output']] = 'Bib Record Fine: ' + mms_id
            elif i == 1:
                if etd_record == None:
                    data[header['output']] = 'ETD Record Not Retrieved: ' + etd_id + ' ' + year + ' ' + xml_id
                else:
                    data[header['output']] = 'ETD Record Fine: ' + etd_id + ' ' + year + ' ' + xml_id
            else:
                data[header['output']] = ''
    else:
        bib_xml = etree.XML(bib_record['anies'][0].encode(encoding="UTF-16"))
        etd_xml = etd_record.pq_xml()
        
        for header in headers:
            header_input = header['input']
            header_output = header['output']
            
            match header_input:
                case 'comments':
                    if opted_out:
                        data[header['output']] = 'Opted-out during 2023 migration'
                        print('\t','OPTOUT')
                    if taken_down:
                        data[header['output']] = 'Requested ProQuest takedown; ' + ('end date on ' + active_embargo if len(active_embargo) > 0 else 'no end date')
                        print('\t','PQ TAKEDOWN')
                    else:
                        data[header['output']] = ''
                #taken from input CSV
                case 'active embargo':
                    data[header['output']] = parse_date_to_iso(active_embargo)
                
                #taken from etd metadata in bag-info.txt
                case 'First-Name' | 'Middle-Name' | 'Last-Name' | 'ProQuest-Email' : 
                    if header['input'] in etd_record.bag.info:
                        data[header['output']] = etd_record.bag.info[header['input']]
                    else:
                        data[header['output']] = ''
                
                #taken from bib data json
                case 'title' | 'mms_id':
                    data[header['output']] = bib_record[header['input']].removesuffix(' /').removesuffix('.')
                
                #taken from bib data json (parsing to ISO date format)
                case 'date_of_publication':
                    raw_string = bib_record[header['input']].removesuffix(' /').removesuffix('.')
                    
                    data[header['output']] = parse_date_to_iso(raw_string)
                    
                #taken from bib data xml; get all subfields
                case '050' | '300' :
                    datafields = bib_xml.xpath("//datafield[@tag=" + header['input'] +"]")
                    data[header['output']] = ''
                    
                    subfield_values = []
                    
                    for datafield in datafields:
                        for subfield in datafield:
                            subfield_values.append(subfield.text)
                    
                    data[header['output']] = ' '.join(subfield_values)
                
                #taken from bib data xml; get all instances, subfield a
                case '336' | '337' | '338' | '650' | '655':
                    subfield_values = get_subfielda_value_from_all(bib_xml, header['input'])
                    data[header['output']] = ' '.join(subfield_values)
                
                #taken from etd's ..._DATA.xml
                
                case 'inst_contact':
                    found_els = etd_xml.find('.//DISS_' + header['input'])
                    if found_els != None:
                        data[header['output']] = ''.join(found_els.itertext())
                        ### needs to go through Department Mapping from a provided XLSX file#########
                    else:
                        data[header['output']] = ''
                case 'para': 
                    found_els = etd_xml.find('.//DISS_' + header['input'])
                    if found_els != None:
                        data[header['output']] = ''.join(found_els.itertext())
                    else:
                        data[header['output']] = ''
                case 'degree':
                    degree_abbrev = etd_xml.find('.//DISS_degree').text or ''
                    data[header['output']] = degree_map[degree_abbrev]
                case 'document_type':
                    degree_abbrev = etd_xml.find('.//DISS_degree').text or ''
                    if degree_abbrev in document_map:
                        data[header['output']] = document_map[degree_abbrev]
                    else:
                        data[header['output']] = ''
                case 'language':
                    lang_code = etd_xml.find('.//DISS_language').text or ''
                    data[header['output']] = Language.make(language=lang_code).display_name()
                case 'advisor':
                    advisors = etd_xml.findall('.//DISS_advisor')
                    
                    data['advisor1'] = ''
                    data['advisor2'] = ''
                    data['advisor3'] = ''
                    
                    for i,advisor in enumerate(advisors):
                        advisor_name = [''] * 3;
                        if advisor.find('.//DISS_fname').text:
                            advisor_name[0] = advisor.find('.//DISS_fname').text or ''
                        if advisor.find('.//DISS_middle').text:
                            advisor_name[1] = advisor.find('.//DISS_middle').text or ''
                        if advisor.find('.//DISS_surname').text:
                            advisor_name[2] = advisor.find('.//DISS_surname').text or ''
                        
                        data['advisor' + str(i+1)] = ' '.join(advisor_name)
                case 'categorization': 
                    categorization = etd_xml.find('.//DISS_cat_code').text or ''
                    data[header['output']] = disciplines[categorization]
                case 'url':
                    if opted_out or (taken_down and len(active_embargo) == 0):
                        data[header['output']] = ''
                    else:
                        data[header['output']] = deposit_pdf(etd_record)
                case 'committee_members': 
                    cmte_members = etd_xml.findall('.//DISS_cmte_member')
                    committee_members = [''] * len(cmte_members)
                    
                    for i,cmte_member in enumerate(cmte_members):
                        member_name = [''] * 3
                        member_name[0] = cmte_member.find('.//DISS_fname').text or ''
                        member_name[1] = cmte_member.find('.//DISS_middle').text or ''
                        member_name[2] = cmte_member.find('.//DISS_surname').text or ''
                        committee_members[i] = ' '.join(member_name)
                    data[header['output']] = ', '.join(committee_members)
                
                # merging from two sources
                case 'keyword,650': 
                    keys = header['input'].split(',')
                    values = []
                    
                    # Retrieve from etd's ..._DATA.xml
                    etd_keywords = []
                    
                    found_els = etd_xml.find('.//DISS_keyword')
                    if found_els is not None and found_els.text is not None:
                        etd_keywords += found_els.text.split(', ')
                    
                    # Retrieve from bib data xml
                    catalog_subjects = get_subfielda_value_from_all(bib_xml, '650')
                    
                    # Merge, removing duplicates
                    merge_list = list(etd_keywords)
                    merge_list.extend(x for x in catalog_subjects if x not in merge_list)
                    
                    data[header['output']] = ','.join(merge_list)
                
                # hard-coded
                case 'inst_name':
                    data[header['output']] = "University at Albany, State University of New York"
                
                #unassigned elsewhere, leave output field blank for now
                case _:
                    data[header['output']] = ''
    return data

#pulls bib data using Alma API
#returns bib JSON or None
def get_bib(mmsid):
    url = 'https://api-na.hosted.exlibrisgroup.com/almaws/v1/bibs/' + mmsid + '?apikey=' + envconfig.api_key
    headers = {'Accept':'application/json'}
    
    response = requests.get(url, headers=headers)
    
    if response.status_code == 200:
        return response.json()
    
    return None

#loads the associated etd object from storage
#returns ETD object or None
def get_etd(etd_id,xml_id,year):
    etd_record = ETD()
    
    if not os.path.isdir(os.path.join(envconfig.storage_directory, year, etd_id)):
        return None;
    else:
        etd_record.load(os.path.join(envconfig.storage_directory, year, etd_id))
        return etd_record

#copies the etd pdf file into a receiving directory visible to BePress
#returns a url string for said file as per the directory's mounting on the Apps Server or empty string
def deposit_pdf(etd_record):
    shutil.copyfile(etd_record.pdf_file,os.path.join(envconfig.working_directory, 'SA_uploads', etd_record.etd_id + '.pdf'))
    
    return 'https://apps.library.albany.edu/sa_uploads/' + etd_record.etd_id + '.pdf'

#parses date from string with various different possible formats and levels of accuracy
#returns a datestring in ISO format (YYYY-MM-DD)
def parse_date_to_iso(input_datestring):
    parsed_date = None
    
    if len(input_datestring) > 0:
        #strip excess spaces
        input_datestring = input_datestring.lstrip()
        
        # attempt to parse as month-day-year (format in takedown spreadsheet)
        try:
            parsed_date = datetime.strptime(input_datestring, '%m-%d-%y').date()
        except:
            pass
        
        #attempt to parse as a month/day/year
        try:
            parsed_date = datetime.strptime(input_datestring, '%m/%d/%Y').date()
        except:
            pass
        #attempt to parse as a month-name year
        if parsed_date is None:
            try:
                parsed_date = datetime.strptime(input_datestring, '%B %Y').date()
            except:
                pass
        #attempt to parse as a season/year
        if parsed_date is None:
            try:
                datestring_array = input_datestring.split()
                season = datestring_array[0]
                year = int(datestring_array[1])
                
                seasons = {
                    'Spring' : 5,
                    'Summer' : 8,
                    'Fall' : 12
                }
                
                parsed_date = datetime(year, seasons[season], 1)
            except:
                pass
        #attempt to parse as a year
        if parsed_date is None:
            try:
                parsed_date = datetime.strptime(input_datestring, '%Y').date()
            except:
                pass
    
    output_datestring = ''
    
    if parsed_date is not None:
        output_datestring = parsed_date.strftime('%Y-%m-%d')
    
    return output_datestring

#retrieves all instances of a datafield
#returns an array of subfield 'a' values from each datafield
def get_subfielda_value_from_all(bib_xml, fieldname):
    datafields = bib_xml.xpath("//datafield[@tag=" + fieldname + "]")
    subfield_values = []
    
    for datafield in datafields:
        subfield_values.append(datafield[0].text.removesuffix('.'))            
    
    return subfield_values

#trigger for main method
if __name__ == "__main__":
    main(sys.argv[1:])